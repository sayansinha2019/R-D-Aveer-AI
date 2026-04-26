"""Compare Project-1 ``Bill of Materials`` sheets by **rollup key** (category, section type,
section, length, grade): **qty**, **weight** (parsed ``lbs`` strings), **length** string parity.

Piecemarks are ignored so shop marks vs generated labels do not block comparison.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

from saddleback_pipeline.bom_relaxed import parse_length_to_inches


def _norm_blank(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip())


def _norm_length(v: Any) -> str:
    """Normalize feet-inches style length for keying."""
    if v is None:
        return ""
    s = str(v).strip().replace("\u2019", "'").replace("\u201d", '"').replace("\u2033", '"')
    return re.sub(r"\s+", " ", s)


def _parse_weight_lbs(val: Any) -> float | None:
    """Parse cells like ``1,975.11 lbs``, ``0 lbs``, ``7.38 lbs``."""
    if val is None:
        return None
    s = str(val).strip().upper().replace(",", "")
    if not s:
        return None
    m = re.search(r"([\d.]+)", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def _normalize_w_designation(section: Any) -> str | None:
    """Map ``24x62``, ``24X62``, ``W24X62`` → ``W24X62``."""
    if section is None:
        return None
    s = str(section).strip().upper().replace(" ", "").replace("×", "X")
    if s.startswith("W") and re.match(r"^W\d+X\d+$", s):
        return s
    m = re.match(r"^(\d+)X(\d+)$", s)
    if m:
        return f"W{m.group(1)}X{m.group(2)}"
    return None


def _rollup(path: Path) -> tuple[dict[tuple[str, ...], dict[str, Any]], list[list[Any]]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bill of Materials"]
    agg: dict[tuple[str, ...], dict[str, Any]] = defaultdict(
        lambda: {"qty": 0, "weight_lb": 0.0, "rows": 0},
    )
    raw_rows: list[list[Any]] = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 19)]
        if not any(v is not None and str(v).strip() for v in row):
            continue
        raw_rows.append(row)
        cat = _norm_blank(row[1])
        sec_type = _norm_blank(row[4])
        section = _norm_blank(row[5])
        length_s = _norm_length(row[6])
        grade = _norm_blank(row[7])
        try:
            q = int(row[3])
        except (TypeError, ValueError):
            q = 0
        wlb = _parse_weight_lbs(row[9])
        key = (cat, sec_type, section, length_s, grade)
        agg[key]["qty"] += q
        if wlb is not None:
            agg[key]["weight_lb"] += wlb
        agg[key]["rows"] += 1
    wb.close()
    return dict(agg), raw_rows


def _rollup_relaxed_w_beams(
    rows: list[list[Any]],
    *,
    length_snap_inches: float = 6.0,
) -> dict[tuple[str, float | None], dict[str, Any]]:
    """Beams + section type W: key = (W designation, snapped length inches). Grade omitted."""
    out: dict[tuple[str, float | None], dict[str, Any]] = defaultdict(
        lambda: {"qty": 0, "weight_lb": 0.0},
    )
    for row in rows:
        cat = _norm_blank(row[1])
        st = _norm_blank(row[4])
        if cat != "Beams" or st != "W":
            continue
        w = _normalize_w_designation(row[5])
        if not w:
            continue
        li = parse_length_to_inches(row[6])
        if li is not None and length_snap_inches > 0:
            li = round(li / length_snap_inches) * length_snap_inches
        key = (w, li)
        try:
            q = int(row[3])
        except (TypeError, ValueError):
            q = 0
        wlb = _parse_weight_lbs(row[9])
        out[key]["qty"] += q
        if wlb is not None:
            out[key]["weight_lb"] += wlb
    return dict(out)


def compare_material_rollups(reference_xlsx: Path, generated_xlsx: Path) -> dict[str, Any]:
    ref_agg, ref_rows = _rollup(reference_xlsx)
    gen_agg, gen_rows = _rollup(generated_xlsx)

    keys_ref = set(ref_agg.keys())
    keys_gen = set(gen_agg.keys())
    both = keys_ref & keys_gen
    only_ref = sorted(keys_ref - keys_gen)
    only_gen = sorted(keys_gen - keys_ref)

    qty_exact = qty_total_ref = qty_total_gen = 0
    qty_abs_err = 0
    wt_err_sum = 0.0
    wt_pairs = 0
    length_match_on_both = 0

    per_key: list[dict[str, Any]] = []
    for k in sorted(both):
        rr, rg = ref_agg[k], gen_agg[k]
        rq, gq = rr["qty"], rg["qty"]
        qty_total_ref += rq
        qty_total_gen += gq
        qty_abs_err += abs(rq - gq)
        if rq == gq:
            qty_exact += 1
        rw, gw = rr["weight_lb"], rg["weight_lb"]
        wt_delta = None
        if rw > 0 or gw > 0:
            wt_pairs += 1
            wt_delta = gw - rw
            wt_err_sum += abs(wt_delta)

        cat, st, sec, ln, gr = k
        per_key.append(
            {
                "category": cat,
                "section_type": st,
                "section": sec,
                "length": ln,
                "grade": gr,
                "reference_qty": rq,
                "generated_qty": gq,
                "qty_delta": gq - rq,
                "reference_weight_lb_sum": round(rw, 2),
                "generated_weight_lb_sum": round(gw, 2),
                "weight_delta_lb": round(wt_delta, 2) if wt_delta is not None else None,
            },
        )
        # length is part of key — if key matches, length string matches by definition
        length_match_on_both += 1

    def mean_abs_pct_err(ref_tot: float, abs_err: float) -> float | None:
        if ref_tot <= 0:
            return None
        return abs_err / ref_tot

    report = {
        "reference_xlsx": str(reference_xlsx.resolve()),
        "generated_xlsx": str(generated_xlsx.resolve()),
        "source_rows": {
            "reference": len(ref_rows),
            "generated": len(gen_rows),
        },
        "distinct_rollup_keys": {
            "reference": len(keys_ref),
            "generated": len(keys_gen),
            "intersection": len(both),
            "only_reference": len(only_ref),
            "only_generated": len(only_gen),
        },
        "key_coverage_recall": len(both) / len(keys_ref) if keys_ref else None,
        "key_coverage_precision": len(both) / len(keys_gen) if keys_gen else None,
        "qty_on_intersection_keys": {
            "reference_total": qty_total_ref,
            "generated_total": qty_total_gen,
            "keys_with_exact_qty_match": qty_exact,
            "keys_compared": len(both),
            "qty_exact_match_rate_on_keys": (qty_exact / len(both)) if both else None,
            "mean_abs_qty_error_per_key": (qty_abs_err / len(both)) if both else None,
            "mean_abs_qty_error_vs_ref_total": mean_abs_pct_err(qty_total_ref, qty_abs_err),
        },
        "weight_on_intersection_keys": {
            "keys_with_any_weight": wt_pairs,
            "mean_abs_weight_error_lb": (wt_err_sum / wt_pairs) if wt_pairs else None,
            "note": "Weight cells parsed from column J; summed per rollup key (same as qty rollup).",
        },
        "length": {
            "definition": "Length is included in rollup key; matching keys imply matching length strings.",
            "matched_keys_imply_length_match_count": length_match_on_both,
        },
        "material_identity": {
            "definition": "Rollup key = Category + Section type + Section + Length + Grade (piecemark excluded).",
        },
        "sample_keys_only_in_reference": [
            {"category": a[0], "section_type": a[1], "section": a[2], "length": a[3], "grade": a[4]}
            for a in only_ref[:25]
        ],
        "sample_keys_only_in_generated": [
            {"category": a[0], "section_type": a[1], "section": a[2], "length": a[3], "grade": a[4]}
            for a in only_gen[:25]
        ],
        "per_key_detail_top_50": per_key[:50],
    }

    ref_rw = _rollup_relaxed_w_beams(ref_rows)
    gen_rw = _rollup_relaxed_w_beams(gen_rows)
    keys_r = set(ref_rw.keys())
    keys_g = set(gen_rw.keys())
    both_w = keys_r & keys_g
    qty_err_w = wt_err_w = 0.0
    qty_match_w = 0
    detail_w = []
    for k in sorted(both_w):
        rq, gq = ref_rw[k]["qty"], gen_rw[k]["qty"]
        if rq == gq:
            qty_match_w += 1
        qty_err_w += abs(rq - gq)
        rw, gw = ref_rw[k]["weight_lb"], gen_rw[k]["weight_lb"]
        wt_err_w += abs(rw - gw)
        wdes, li = k
        detail_w.append(
            {
                "w_section": wdes,
                "length_inches_snapped_6in": li,
                "reference_qty": rq,
                "generated_qty": gq,
                "qty_delta": gq - rq,
                "reference_weight_lb": round(rw, 2),
                "generated_weight_lb": round(gw, 2),
                "weight_abs_delta_lb": round(abs(rw - gw), 2),
            },
        )

    report["relaxed_w_beam_keys"] = {
        "description": "Category=Beams, Section type=W; key = normalized W designation + length (in) snapped to 6\"; grade ignored.",
        "distinct_keys_reference": len(keys_r),
        "distinct_keys_generated": len(keys_g),
        "keys_in_both": len(both_w),
        "qty_exact_match_on_shared_keys": qty_match_w,
        "qty_match_rate_on_shared": (qty_match_w / len(both_w)) if both_w else None,
        "mean_abs_qty_error_on_shared": (qty_err_w / len(both_w)) if both_w else None,
        "mean_abs_weight_lb_error_on_shared": (wt_err_w / len(both_w)) if both_w else None,
        "shared_key_detail": detail_w[:80],
    }

    return report


def main() -> int:
    p = argparse.ArgumentParser(
        description="Qty / weight / length / material-key accuracy vs reference Project-1 BOM.",
    )
    p.add_argument("--reference", type=Path, required=True)
    p.add_argument("--generated", type=Path, required=True)
    p.add_argument("--out-json", type=Path, default=None)
    args = p.parse_args()
    report = compare_material_rollups(
        args.reference.expanduser().resolve(),
        args.generated.expanduser().resolve(),
    )
    text = json.dumps(report, indent=2, ensure_ascii=False)
    print(text)
    if args.out_json:
        outp = args.out_json.expanduser().resolve()
        outp.parent.mkdir(parents=True, exist_ok=True)
        outp.write_text(text, encoding="utf-8")
        print(f"Wrote {outp}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
