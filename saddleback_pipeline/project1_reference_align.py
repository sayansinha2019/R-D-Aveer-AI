"""Align takeoff entities to a **reference** Project-1 ``Bill of Materials`` XLSX.

Use when you have an authoritative shop/fabricator BOM for the same job (or a prior run that
matches the PDFs). This fills **weights**, **grades**, and optionally **piecemarks** on takeoff
``data[]`` rows when ``(category, section display, length)`` matches within tolerance — improving
parity before nominal AISC lb/ft and CSV overrides run.

Priority (recommended pipeline order)
-----------------------------------
1. This module (reference BOM rows; typically fill-null-only).
2. ``steel_weight_enrichment`` nominal lb/ft × length for still-missing rolled shapes.
3. Weight override CSV for HSS / plates / odd marks (can use ``overwrite_existing`` there if needed).

This does **not** replace a Tekla/SDS2 model export as full truth; it narrows the gap for PDF-first
workflows when a matching BOM exists.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

from saddleback_pipeline.bom_relaxed import material_match_key, parse_length_to_inches
from saddleback_pipeline.project1_bom_export import _classify_row, _section_display, _strip
from saddleback_pipeline.steel_weight_enrichment import normalize_steel_section_callout


def _parse_weight_lbs(val: Any) -> float | None:
    if val is None:
        return None
    s = str(val).strip().upper().replace(",", "")
    if not s:
        return None
    m = re.search(r"([\d.]+)", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def _norm_length_cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip().replace("\u2019", "'").replace("\u201d", '"').replace("\u2033", '"')
    return re.sub(r"\s+", " ", s)


def load_project1_bom_reference_rows(path: Path) -> list[dict[str, Any]]:
    path = path.expanduser().resolve()
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bill of Materials"]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(r, c).value for c in range(1, 19)]
        if not any(v is not None and str(v).strip() for v in cells):
            continue
        cat = _strip(cells[1])
        sec_type = _strip(cells[4])
        section = _strip(cells[5])
        length_s = _norm_length_cell(cells[6])
        grade = _strip(cells[7])
        try:
            qty = int(cells[3] or 0)
        except (TypeError, ValueError):
            qty = 0
        wlb = _parse_weight_lbs(cells[9])
        piece = _strip(cells[2])
        rows.append(
            {
                "category": cat,
                "section_type": sec_type,
                "section": section,
                "length_str": length_s,
                "length_in": parse_length_to_inches(length_s),
                "grade": grade,
                "qty": max(1, qty) if qty else 1,
                "weight_total": wlb,
                "piecemark": piece,
                "main": cells[0],
            },
        )
    wb.close()
    return rows


def _display_section_for_entity(ent: dict[str, Any]) -> tuple[str, str, str]:
    cat, sec_type, _, _ = _classify_row(ent)
    disp = _section_display(ent, sec_type)
    raw = _strip(ent.get("section"))
    return cat, sec_type, disp or raw


def _section_keys_for_match(ent: dict[str, Any], sec_type: str) -> set[str]:
    _, _, disp = _display_section_for_entity(ent)
    keys: set[str] = set()
    if disp:
        keys.add(material_match_key(disp))
    raw = _strip(ent.get("section"))
    if raw:
        keys.add(material_match_key(raw))
        keys.add(material_match_key(normalize_steel_section_callout(raw)))
    if sec_type == "W" and raw:
        m = re.search(r"(?:^W)?\s*(\d+)\s*[Xx]\s*(\d+)", raw.replace(" ", ""), re.I)
        if m:
            keys.add(material_match_key(f"{m.group(1)}x{m.group(2)}"))
            keys.add(material_match_key(f"W{m.group(1)}X{m.group(2)}"))
    return {k for k in keys if k}


def _ref_section_keys(ref: dict[str, Any]) -> set[str]:
    st = ref.get("section_type") or ""
    sec = _strip(ref.get("section"))
    keys: set[str] = {material_match_key(sec)}
    if st == "W":
        m = re.match(r"^(\d+)x(\d+)$", sec.replace(" ", "").replace("×", "x"), re.I)
        if m:
            keys.add(material_match_key(f"W{m.group(1)}X{m.group(2)}"))
    keys.add(material_match_key(normalize_steel_section_callout(sec)))
    return {k for k in keys if k}


def _sections_match(ent: dict[str, Any], ref: dict[str, Any], classified_stype: str) -> bool:
    if (ref.get("section_type") or "").strip() != (classified_stype or "").strip():
        return False
    a = _section_keys_for_match(ent, classified_stype)
    b = _ref_section_keys(ref)
    return bool(a & b)


def _pick_reference_row(
    ent: dict[str, Any],
    ref_rows: list[dict[str, Any]],
    *,
    tol_inches: float,
    allow_section_only_when_length_missing: bool = False,
    fallback_nearest_if_no_within_tol: bool = False,
) -> dict[str, Any] | None:
    cat, st, _ = _display_section_for_entity(ent)
    gen_li = parse_length_to_inches(ent.get("length"))
    best_within_tol: tuple[float, dict[str, Any]] | None = None
    best_any: tuple[float, dict[str, Any]] | None = None
    section_only_pool: list[dict[str, Any]] = []

    for ref in ref_rows:
        if ref.get("category") != cat:
            continue
        if not _sections_match(ent, ref, st):
            continue
        section_only_pool.append(ref)
        rli = ref.get("length_in")
        dist: float
        if gen_li is not None and rli is not None:
            dist = abs(float(rli) - float(gen_li))
            if best_any is None or dist < best_any[0]:
                best_any = (dist, ref)
            if tol_inches >= 0 and dist > tol_inches:
                continue
        elif gen_li is None and rli is None:
            dist = 0.0
        elif gen_li is None and allow_section_only_when_length_missing:
            # Accept section-only match when generated length is missing.
            dist = 0.0
        else:
            continue

        if best_within_tol is None or dist < best_within_tol[0]:
            best_within_tol = (dist, ref)

    if best_within_tol is not None:
        return best_within_tol[1]
    if fallback_nearest_if_no_within_tol and best_any is not None:
        return best_any[1]
    if allow_section_only_when_length_missing and gen_li is None and section_only_pool:
        # Prefer rows with explicit length and larger qty footprint.
        pool = sorted(
            section_only_pool,
            key=lambda r: (
                0 if r.get("length_in") is not None else 1,
                -int(r.get("qty") or 1),
            ),
        )
        return pool[0]
    return None


def align_takeoff_payload(
    payload: dict[str, Any],
    *,
    reference_xlsx: Path,
    tol_inches: float = 6.0,
    fill_piecemarks: bool = False,
    only_fill_empty_weight: bool = True,
    only_fill_empty_grade: bool = True,
    fill_length_from_reference: bool = False,
    allow_section_only_when_length_missing: bool = False,
    fallback_nearest_if_no_within_tol: bool = False,
    categories: set[str] | None = None,
) -> dict[str, Any]:
    ref_rows = load_project1_bom_reference_rows(Path(reference_xlsx))
    stats = {
        "reference_rows_loaded": len(ref_rows),
        "entities_matched": 0,
        "weight_filled": 0,
        "grade_filled": 0,
        "length_filled": 0,
        "piecemark_filled": 0,
        "skipped_no_match": 0,
        "skipped_existing_weight": 0,
        "skipped_existing_grade": 0,
    }

    data = payload.get("data")
    if not isinstance(data, list):
        payload.setdefault("meta", {})
        if isinstance(payload["meta"], dict):
            payload["meta"]["project1_reference_align"] = stats
        return stats

    for ent in data:
        if not isinstance(ent, dict):
            continue
        ent_cat, _ent_st, _ent_sec = _display_section_for_entity(ent)
        if categories is not None and ent_cat not in categories:
            continue
        ref = _pick_reference_row(
            ent,
            ref_rows,
            tol_inches=tol_inches,
            allow_section_only_when_length_missing=allow_section_only_when_length_missing,
            fallback_nearest_if_no_within_tol=fallback_nearest_if_no_within_tol,
        )
        if ref is None:
            stats["skipped_no_match"] += 1
            continue
        stats["entities_matched"] += 1

        if fill_length_from_reference:
            rlen = _norm_length_cell(ref.get("length_str"))
            if rlen:
                if _norm_length_cell(ent.get("length")) != rlen:
                    ent["length"] = rlen
                    stats["length_filled"] += 1

        w0 = ent.get("weight")
        has_w = False
        if w0 is not None:
            try:
                float(w0)
                has_w = True
            except (TypeError, ValueError):
                pass
        wt = ref.get("weight_total")
        may_write_weight = wt is not None and (not only_fill_empty_weight or not has_w)
        if may_write_weight:
            rq = int(ref.get("qty") or 1)
            per_piece = float(wt) / max(1, rq)
            try:
                q = int(ent.get("quantity") or 1)
            except (TypeError, ValueError):
                q = 1
            ent["weight"] = round(per_piece * max(1, q), 2)
            ent["weight_source"] = "reference_project1_bom"
            stats["weight_filled"] += 1
        elif has_w and only_fill_empty_weight:
            stats["skipped_existing_weight"] += 1

        gr = _strip(ref.get("grade"))
        if gr:
            cur_mat = _strip(ent.get("material"))
            if not cur_mat or not only_fill_empty_grade:
                ent["material"] = gr
                stats["grade_filled"] += 1
            else:
                stats["skipped_existing_grade"] += 1

        pm = _strip(ent.get("piece_mark"))
        rpm = ref.get("piecemark") or ""
        if fill_piecemarks and rpm and not pm:
            ent["piece_mark"] = rpm
            stats["piecemark_filled"] += 1

    # material_summary: loose match on material string + length
    bom = payload.get("material_summary")
    if isinstance(bom, list) and ref_rows:
        for row in bom:
            if not isinstance(row, dict):
                continue
            mat = row.get("material")
            mkey = material_match_key(mat)
            if not mkey:
                continue
            rli = parse_length_to_inches(row.get("length"))
            best_r: dict[str, Any] | None = None
            best_d = tol_inches + 1.0
            for ref in ref_rows:
                if mkey not in _ref_section_keys(ref):
                    continue
                ref_li = ref.get("length_in")
                if rli is not None and ref_li is not None:
                    d = abs(float(rli) - float(ref_li))
                    if d <= tol_inches and d < best_d:
                        best_d = d
                        best_r = ref
                elif rli is None and ref_li is None:
                    best_r = ref
                    best_d = 0.0
                    break
            if best_r is None:
                continue
            rw = row.get("weight")
            has_rw = rw is not None
            try:
                if has_rw:
                    float(rw)
            except (TypeError, ValueError):
                has_rw = False
            wtot = best_r.get("weight_total")
            if wtot is None:
                continue
            if only_fill_empty_weight and has_rw:
                continue
            qty = int(row.get("qty") or 1)
            rq = int(best_r.get("qty") or 1)
            per_piece = float(wtot) / max(1, rq)
            row["weight"] = round(per_piece * max(1, qty), 2)
            row["weight_source"] = "reference_project1_bom"

    payload.setdefault("meta", {})
    if isinstance(payload["meta"], dict):
        payload["meta"]["project1_reference_align"] = {**stats, "tolerance_inches": tol_inches}
    return stats


def align_takeoff_json_file(
    path: Path,
    *,
    reference_xlsx: Path,
    tol_inches: float = 6.0,
    fill_piecemarks: bool = False,
    only_fill_empty_weight: bool = True,
    only_fill_empty_grade: bool = True,
    fill_length_from_reference: bool = False,
    allow_section_only_when_length_missing: bool = False,
    fallback_nearest_if_no_within_tol: bool = False,
    categories: set[str] | None = None,
) -> dict[str, Any]:
    path = path.expanduser().resolve()
    payload = json.loads(path.read_text(encoding="utf-8"))
    stats = align_takeoff_payload(
        payload,
        reference_xlsx=reference_xlsx,
        tol_inches=tol_inches,
        fill_piecemarks=fill_piecemarks,
        only_fill_empty_weight=only_fill_empty_weight,
        only_fill_empty_grade=only_fill_empty_grade,
        fill_length_from_reference=fill_length_from_reference,
        allow_section_only_when_length_missing=allow_section_only_when_length_missing,
        fallback_nearest_if_no_within_tol=fallback_nearest_if_no_within_tol,
        categories=categories,
    )
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return stats


def main() -> int:
    p = argparse.ArgumentParser(description="Align takeoff JSON to Project-1 reference BOM xlsx.")
    p.add_argument("--json", type=Path, required=True)
    p.add_argument("--reference-xlsx", type=Path, required=True)
    p.add_argument("--tolerance-inches", type=float, default=6.0)
    p.add_argument("--fill-piecemarks", action="store_true")
    p.add_argument("--fill-length-from-reference", action="store_true")
    p.add_argument(
        "--allow-section-only-when-length-missing",
        action="store_true",
        help="Permit match by category+section when generated length is missing.",
    )
    p.add_argument(
        "--fallback-nearest-if-no-within-tol",
        action="store_true",
        help="If no match within tolerance, use nearest section-matched reference length.",
    )
    p.add_argument(
        "--categories",
        nargs="*",
        default=[],
        help='Optional filter, e.g. --categories Beams Columns',
    )
    args = p.parse_args()
    st = align_takeoff_json_file(
        args.json,
        reference_xlsx=args.reference_xlsx,
        tol_inches=args.tolerance_inches,
        fill_piecemarks=args.fill_piecemarks,
        fill_length_from_reference=args.fill_length_from_reference,
        allow_section_only_when_length_missing=args.allow_section_only_when_length_missing,
        fallback_nearest_if_no_within_tol=args.fallback_nearest_if_no_within_tol,
        categories=set(args.categories) if args.categories else None,
    )
    print(json.dumps(st, indent=2))
    print(f"Updated: {args.json}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
