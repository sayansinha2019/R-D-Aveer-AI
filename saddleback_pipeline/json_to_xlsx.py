"""Export takeoff_output.json (root key ``data``) to a structured Excel workbook."""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def _scalar_or_join(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(str(x) for x in value)
    return str(value)


# Primary sheet: one row per entity, human-readable columns
_COLUMNS: list[tuple[str, str]] = [
    ("entity_id", "Entity ID"),
    ("parent_group", "Parent Group"),
    ("element_type", "Element Type"),
    ("section", "Section / Size"),
    ("material", "Material"),
    ("length", "Length"),
    ("plate_length", "Plate Length"),
    ("plate_width", "Plate Width"),
    ("plate_thickness", "Plate Thickness"),
    ("quantity", "Qty"),
    ("total_group_quantity", "Group Qty"),
    ("piece_mark", "Piece Mark"),
    ("calculation_steps", "Calculation Steps"),
    ("quantity_reasoning", "Quantity Reasoning"),
    ("source_reference", "Source Reference"),
]

# Fabrication BOM sheet (aligned with typical Material Summary Excel)
_BOM_COLUMNS: list[tuple[str, str]] = [
    ("qty", "Qty"),
    ("material", "Material"),
    ("length", "Length"),
    ("pcmk", "Pcmk"),
    ("weight", "Weight"),
    ("grade", "Grade"),
]


def export_takeoff_json_to_xlsx(
    input_json: Path,
    output_xlsx: Path,
) -> tuple[int, int, int]:
    input_json = input_json.expanduser().resolve()
    if not input_json.is_file():
        raise FileNotFoundError(f"JSON not found: {input_json}")

    payload = json.loads(input_json.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = payload.get("data") or []
    if not isinstance(rows, list):
        raise ValueError('Expected JSON object with a "data" array')
    bom_rows: list[dict[str, Any]] = payload.get("material_summary") or []
    if bom_rows and not isinstance(bom_rows, list):
        bom_rows = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Takeoff"

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, (_key, title) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = wrap

    for r_idx, row in enumerate(rows, start=2):
        if not isinstance(row, dict):
            continue
        for col_idx, (key, _title) in enumerate(_COLUMNS, start=1):
            raw = row.get(key)
            val = _scalar_or_join(raw)
            c = ws.cell(row=r_idx, column=col_idx, value=val)
            c.alignment = wrap

    ws.freeze_panes = "A2"
    # Reasonable default widths (characters)
    widths = {
        "A": 12,
        "B": 16,
        "C": 14,
        "D": 28,
        "E": 14,
        "F": 12,
        "G": 12,
        "H": 12,
        "I": 14,
        "J": 6,
        "K": 10,
        "L": 12,
        "M": 48,
        "N": 40,
        "O": 36,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    if bom_rows:
        ws2 = wb.create_sheet("Material Summary", 1)
        for col_idx, (_key, title) in enumerate(_BOM_COLUMNS, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for r_idx, brow in enumerate(bom_rows, start=2):
            if not isinstance(brow, dict):
                continue
            for col_idx, (key, _title) in enumerate(_BOM_COLUMNS, start=1):
                raw = brow.get(key)
                if key == "weight" and raw is not None:
                    val: Any = raw
                else:
                    val = _scalar_or_join(raw)
                c = ws2.cell(row=r_idx, column=col_idx, value=val)
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.freeze_panes = "A2"
        for i, letter in enumerate("ABCDEF", start=1):
            ws2.column_dimensions[letter].width = [8, 22, 18, 14, 10, 14][i - 1]

    n_scale = 0
    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    ds = meta.get("drawing_scales") if isinstance(meta.get("drawing_scales"), dict) else {}
    scale_pages = ds.get("pages") if isinstance(ds.get("pages"), list) else []
    if scale_pages:
        ws3 = wb.create_sheet("Drawing scales", 2)
        sh = ("Page", "Raw", "feet_per_drawing_inch", "drawing_inches", "real_feet", "kind")
        for col_idx, title in enumerate(sh, start=1):
            c = ws3.cell(row=1, column=col_idx, value=title)
            c.font = header_font
        r = 2
        for pg in scale_pages:
            if not isinstance(pg, dict):
                continue
            pno = pg.get("page")
            for s in pg.get("scales") or []:
                if not isinstance(s, dict):
                    continue
                ws3.cell(row=r, column=1, value=pno)
                ws3.cell(row=r, column=2, value=s.get("raw"))
                ws3.cell(row=r, column=3, value=s.get("feet_per_drawing_inch"))
                ws3.cell(row=r, column=4, value=s.get("drawing_inches"))
                ws3.cell(row=r, column=5, value=s.get("real_feet"))
                ws3.cell(row=r, column=6, value=s.get("kind"))
                r += 1
                n_scale += 1
        ws3.freeze_panes = "A2"
        for i, letter in enumerate("ABCDEF", start=1):
            ws3.column_dimensions[letter].width = [6, 40, 18, 14, 12, 18][i - 1]

    output_xlsx = output_xlsx.expanduser().resolve()
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return len(rows), len(bom_rows), n_scale


def main() -> int:
    import os

    from dotenv import load_dotenv

    load_dotenv(".env", override=False)
    in_path = (os.getenv("INPUT_JSON", "") or "").strip() or "takeoff_output.json"
    out_path = (os.getenv("OUTPUT_XLSX", "") or "").strip() or "takeoff_output.xlsx"

    n_ent, n_bom, n_sc = export_takeoff_json_to_xlsx(Path(in_path), Path(out_path))
    print(
        f"Wrote {out_path} (Takeoff: {n_ent} rows; Material Summary: {n_bom} rows; "
        f"Drawing scales: {n_sc} rows)",
        file=sys.stderr,
    )
    print(out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
