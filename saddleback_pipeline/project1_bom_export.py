"""Export takeoff JSON to Project-1 style ``Bill of Materials`` XLSX (18 columns)."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Matches reference: ``Project 1/Project-1 - BOM - April 19, 2026.xlsx``
_BOM_HEADERS: list[str] = [
    "Main",
    "Category",
    "Piecemark",
    "Qty",
    "Section Type",
    "Section",
    "Length",
    "Grade",
    "Labor Code",
    "Weight",
    "Camber",
    "Cope",
    "Hole",
    "Weld Stud",
    "Status",
    "Sequence",
    "DCR Left",
    "DCR Right",
]

_TEMPLATE_COL_WIDTHS: list[float] = [
    10.83203125,
    12.83203125,
    11.83203125,
    10.83203125,
    14.83203125,
    18.83203125,
    12.83203125,
    10.83203125,
    12.83203125,
    14.83203125,
    10.83203125,
    10.83203125,
    10.83203125,
    11.83203125,
    13.83203125,
    10.83203125,
    21.83203125,
    21.83203125,
]


def _strip(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _parse_camber(section: str) -> str | None:
    m = re.search(r"C\s*=\s*([^)\s]+)", section, re.I)
    if m:
        return m.group(1).strip().strip('"').strip("'")
    return None


def _strip_camber_notation(section: str) -> str:
    return re.sub(r"\s*C\s*=\s*[^)\s]+", "", section, flags=re.I).strip()


def _norm_grade(material: str) -> str | None:
    if not material:
        return None
    if "#" in material or "LW" in material or "SW" in material:
        return material.strip()
    t = material.upper()
    if "A992" in t:
        return "A992"
    if "A992" in material:
        return "A992"
    if "A500" in t:
        m = re.search(r"A500[^A-Z0-9]*([A-Z0-9.]+)", t)
        if m:
            return f"A500-GR.{m.group(1).replace('GR.', '').replace('GR', '')}"
        return "A500-GR.B"
    if "A36" in t:
        return "A36"
    if "A325" in t:
        if "N" in t or "N" in material:
            return "A325N"
        return "A325"
    if "F1554" in t:
        return "F1554-GR.36" if "36" in t else material.strip()
    # Fallback: short last token if looks like a grade
    return material.strip() if len(material) < 24 else None


def _format_weight_lbs(val: Any) -> str:
    if val is None or val == "":
        return "0 lbs"
    if isinstance(val, (int, float)):
        return f"{float(val):,.2f} lbs"
    s = str(val).strip()
    if "lb" in s.lower():
        return s
    try:
        return f"{float(s.replace(',', '')):,.2f} lbs"
    except ValueError:
        return s if s else "0 lbs"


def _w_section_parts(section_raw: str) -> tuple[str, str] | None:
    s = section_raw.upper().replace(" ", "")
    m = re.match(r"^W(\d+)X(\d+)([A-Z])?$", s)
    if m:
        return "W", f"{m.group(1)}x{m.group(2)}"
    return None


def _hss_section_parts(section_raw: str) -> tuple[str, str] | None:
    s = section_raw.upper().replace(" ", "")
    if not s.startswith("HSS"):
        return None
    return "HSS", re.sub(r"X", "x", section_raw.strip(), flags=re.I)


def _classify_row(entity: dict[str, Any]) -> tuple[str, str, str | None, bool]:
    """Return (category, section_type, labor_code, is_main_member)."""
    et = _strip(entity.get("element_type")).lower()
    pg = _strip(entity.get("parent_group")).lower()
    sec_raw = _strip(entity.get("section"))
    sec = _strip_camber_notation(sec_raw) or sec_raw

    if "footing" in et or pg == "foundations":
        return "Foundations", "PL", "Y", False

    if "stud" in et or "weld stud" in et or sec.upper().startswith("WS"):
        return "Weld Studs", "WS", None, False
    if "anchor" in et or sec.upper().startswith("AB") or "ANCHOR" in sec.upper():
        return "Anchors", "AB", None, False
    if "bolt" in et or "bolt" in pg:
        return "Bolts", "HS", None, False
    if "plate" in et or ("base" in et and "plate" in et) or et.replace(" ", "") == "baseplate":
        return "Plates", "PL", "Y", False
    if sec.upper().startswith("PL"):
        return "Plates", "PL", "Y", False
    if "clip" in et or "angle" in et or "l-shape" in et or sec.upper().startswith("L"):
        return "Clips", "L", "A", False
    if "column" in et or "column" in pg:
        wp = _w_section_parts(sec)
        if wp:
            return "Columns", wp[0], "B", True
        hp = _hss_section_parts(sec)
        if hp:
            return "Columns", hp[0], "B", True
        return "Columns", "W", "B", True
    if "beam" in et or "rafter" in et or "girder" in et or "joist" in et:
        wp = _w_section_parts(sec)
        if wp:
            return "Beams", wp[0], "B", True
        hp = _hss_section_parts(sec)
        if hp:
            return "Beams", hp[0], "B", True
        return "Beams", "W", "B", True
    if sec.upper().startswith("W"):
        wp = _w_section_parts(sec)
        if wp:
            return "Beams", wp[0], "B", True
    if sec.upper().startswith("HSS"):
        hp = _hss_section_parts(sec)
        if hp:
            return "Beams", hp[0], "B", True
    return "Beams", "W", "B", True


def _section_display(entity: dict[str, Any], section_type: str) -> str:
    sec_raw = _strip(entity.get("section"))
    sec = _strip_camber_notation(sec_raw) or sec_raw
    if _strip(entity.get("element_type")).lower() == "footing":
        return sec_raw or sec
    if section_type == "W":
        wp = _w_section_parts(sec)
        if wp:
            return wp[1]
        m = re.search(r"W\s*(\d+)\s*[Xx]\s*(\d+)", sec, re.I)
        if m:
            return f"{m.group(1)}x{m.group(2)}"
    if section_type == "HSS":
        hp = _hss_section_parts(sec)
        if hp:
            return hp[1]
    if section_type in {"PL", "HS", "L", "AB", "WS"}:
        return sec
    return sec


def _qty(entity: dict[str, Any]) -> int:
    q = entity.get("quantity")
    try:
        return int(q) if q is not None else 1
    except (TypeError, ValueError):
        return 1


def entity_to_bom_row(entity: dict[str, Any], sequence: int) -> list[Any]:
    category, section_type, labor, is_main = _classify_row(entity)
    sec_raw = _strip(entity.get("section"))
    camber = _parse_camber(sec_raw)
    piece = _strip(entity.get("piece_mark")) or _strip(entity.get("entity_id")) or "UNMARKED"
    grade = _norm_grade(_strip(entity.get("material")))
    length = _strip(entity.get("length")) or None
    if not length:
        plen = _strip(entity.get("plate_length"))
        pwid = _strip(entity.get("plate_width"))
        pth = _strip(entity.get("plate_thickness"))
        if plen and pwid:
            length = f"{plen} x {pwid}" if not pth else f"{pth} x {pwid}"
    section_disp = _section_display(entity, section_type)
    weight = _format_weight_lbs(entity.get("weight"))

    return [
        bool(is_main),
        category,
        piece,
        _qty(entity),
        section_type,
        section_disp or None,
        length,
        grade,
        labor,
        weight,
        camber,
        None,
        None,
        None,
        "Not Set",
        sequence,
        None,
        None,
    ]


def _write_bom_xlsx_for_entities(
    rows_in: list[dict[str, Any]],
    output_xlsx: Path,
    *,
    template_xlsx: Path | None = None,
) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of Materials"
    bold = Font(bold=True)
    for col, title in enumerate(_BOM_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = bold
        c.alignment = Alignment(vertical="top", wrap_text=True)

    from openpyxl.utils import get_column_letter

    row_idx = 2
    seq = 1
    for ent in rows_in:
        if not isinstance(ent, dict):
            continue
        vals = entity_to_bom_row(ent, seq)
        seq += 1
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row_idx += 1

    for i, w in enumerate(_TEMPLATE_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    output_xlsx = output_xlsx.expanduser().resolve()
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    _ = template_xlsx  # reserved for future column style / validation parity

    return row_idx - 2


def export_takeoff_json_to_project1_bom_xlsx(
    input_json: Path,
    output_xlsx: Path,
    *,
    template_xlsx: Path | None = None,
) -> int:
    input_json = input_json.expanduser().resolve()
    payload = json.loads(input_json.read_text(encoding="utf-8"))
    rows_in: list[dict[str, Any]] = payload.get("data") or []
    if not isinstance(rows_in, list):
        raise ValueError('Expected JSON with a "data" array')
    return _write_bom_xlsx_for_entities(rows_in, output_xlsx, template_xlsx=template_xlsx)


def export_merged_takeoff_jsons_to_project1_bom_xlsx(
    input_jsons: list[Path],
    output_xlsx: Path,
    *,
    template_xlsx: Path | None = None,
) -> int:
    """Concatenate ``data`` from several takeoff JSON files (order preserved) → one BOM sheet."""
    merged: list[dict[str, Any]] = []
    for p in input_jsons:
        p = p.expanduser().resolve()
        payload = json.loads(p.read_text(encoding="utf-8"))
        chunk = payload.get("data") or []
        if not isinstance(chunk, list):
            raise ValueError(f'Expected "data" array in {p}')
        merged.extend(e for e in chunk if isinstance(e, dict))
    return _write_bom_xlsx_for_entities(merged, output_xlsx, template_xlsx=template_xlsx)


def main() -> int:
    p = argparse.ArgumentParser(
        description="Export takeoff JSON to Project-1 BOM XLSX. "
        "Pass multiple --json paths to merge ``data`` into one workbook.",
    )
    p.add_argument(
        "--json",
        required=True,
        nargs="+",
        type=Path,
        help="One or more takeoff_output.json paths (merged in list order when multiple).",
    )
    p.add_argument("--out", required=True, type=Path, help="Output .xlsx path")
    p.add_argument(
        "--template",
        type=Path,
        default=None,
        help="Optional reference BOM workbook (reserved for style parity).",
    )
    args = p.parse_args()
    if len(args.json) == 1:
        n = export_takeoff_json_to_project1_bom_xlsx(args.json[0], args.out, template_xlsx=args.template)
    else:
        n = export_merged_takeoff_jsons_to_project1_bom_xlsx(
            list(args.json),
            args.out,
            template_xlsx=args.template,
        )
    print(f"Wrote {args.out} ({n} BOM rows)", file=sys.stderr)
    print(args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
