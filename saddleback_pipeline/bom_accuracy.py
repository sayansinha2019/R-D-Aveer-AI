"""Compare ``material_summary`` in takeoff JSON to a reference Material Summary xlsx."""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from dataclasses import dataclass
from typing import Any

import openpyxl


def _norm_len(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t.replace("\u2019", "'").replace("\u201d", '"')


def _norm_mat(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    return t.upper()


def _norm_pcmk(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())


def _norm_grade(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).upper()


@dataclass
class MaterialRollup:
    """Totals for one normalized material string (length ignored)."""

    qty: int
    weight_sum: float | None
    pcmks: frozenset[str]
    grades: frozenset[str]


def load_reference_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    rows: list[dict[str, Any]] = []
    for r in raw:
        if not r or all(v is None for v in r[:6]):
            continue
        cells = (list(r) + [None] * 6)[:6]
        qty, mat, ln, pcmk, weight, grade = cells
        try:
            q = int(qty)
        except (TypeError, ValueError):
            continue
        rows.append(
            {
                "qty": q,
                "material": mat,
                "length": ln,
                "pcmk": pcmk,
                "weight": weight,
                "grade": grade,
            }
        )
    return rows


def rollup_by_material(rows: list[dict[str, Any]]) -> dict[str, MaterialRollup]:
    """Sum qty and weight per normalized material; collect unique pcmk / grade strings."""
    acc: dict[str, dict[str, Any]] = {}
    for row in rows:
        m = _norm_mat(row.get("material"))
        if not m:
            continue
        if m not in acc:
            acc[m] = {
                "qty": 0,
                "weights": [],
                "pcmks": set(),
                "grades": set(),
            }
        acc[m]["qty"] += int(row["qty"])
        w = row.get("weight")
        if isinstance(w, (int, float)) and not isinstance(w, bool):
            acc[m]["weights"].append(float(w))
        p = row.get("pcmk")
        if p is not None and str(p).strip():
            acc[m]["pcmks"].add(_norm_pcmk(p))
        g = row.get("grade")
        if g is not None and str(g).strip():
            acc[m]["grades"].add(_norm_grade(g))

    out: dict[str, MaterialRollup] = {}
    for m, d in acc.items():
        ws = d["weights"]
        wsum: float | None = sum(ws) if ws else None
        out[m] = MaterialRollup(
            qty=d["qty"],
            weight_sum=wsum,
            pcmks=frozenset(d["pcmks"]) if d["pcmks"] else frozenset(),
            grades=frozenset(d["grades"]) if d["grades"] else frozenset(),
        )
    return out


def load_generated_bom_rows(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    bom = payload.get("material_summary")
    if not bom or not isinstance(bom, list):
        return []
    rows: list[dict[str, Any]] = []
    for row in bom:
        if not isinstance(row, dict):
            continue
        try:
            q = int(row.get("qty"))
        except (TypeError, ValueError):
            continue
        rows.append(
            {
                "qty": q,
                "material": row.get("material"),
                "length": row.get("length"),
                "pcmk": row.get("pcmk"),
                "weight": row.get("weight"),
                "grade": row.get("grade"),
            }
        )
    return rows


def compare_columns_by_material(
    ref_roll: dict[str, MaterialRollup],
    gen_roll: dict[str, MaterialRollup],
) -> dict[str, Any]:
    ref_mats = set(ref_roll)
    gen_mats = set(gen_roll)
    both = ref_mats & gen_mats
    only_ref = ref_mats - gen_mats
    only_gen = gen_mats - ref_mats

    n_both = len(both)
    qty_ok = weight_ok = grade_ok = pcmk_ok = 0
    weight_both_numeric = 0
    qty_mismatches: list[tuple[str, int, int]] = []
    weight_mismatches: list[tuple[str, float | None, float | None]] = []
    grade_mismatches: list[tuple[str, frozenset, frozenset]] = []
    pcmk_mismatches: list[tuple[str, frozenset, frozenset]] = []

    for m in both:
        r, g = ref_roll[m], gen_roll[m]
        if r.qty == g.qty:
            qty_ok += 1
        else:
            qty_mismatches.append((m, r.qty, g.qty))

        rw, gw = r.weight_sum, g.weight_sum
        if rw is not None and gw is not None:
            weight_both_numeric += 1
            if abs(rw - gw) < 0.51:
                weight_ok += 1
            else:
                weight_mismatches.append((m, rw, gw))
        elif rw is None and gw is None:
            pass
        else:
            weight_mismatches.append((m, rw, gw))

        if r.grades == g.grades:
            grade_ok += 1
        else:
            grade_mismatches.append((m, r.grades, g.grades))

        if r.pcmks == g.pcmks:
            pcmk_ok += 1
        else:
            pcmk_mismatches.append((m, r.pcmks, g.pcmks))

    def rate(ok: int) -> float:
        return ok / n_both if n_both else 0.0

    return {
        "materials_reference": len(ref_mats),
        "materials_generated": len(gen_mats),
        "materials_in_both": n_both,
        "only_reference_materials": sorted(only_ref),
        "only_generated_materials": sorted(only_gen),
        "material_recall": len(both) / len(ref_mats) if ref_mats else 0.0,
        "material_precision": len(both) / len(gen_mats) if gen_mats else 0.0,
        "qty_match_rate_on_both": rate(qty_ok),
        "qty_matches": qty_ok,
        "qty_mismatches": qty_mismatches,
        "weight_match_rate_both_numeric": (
            weight_ok / weight_both_numeric if weight_both_numeric else None
        ),
        "weight_materials_both_numeric": weight_both_numeric,
        "weight_ok_count": weight_ok,
        "weight_mismatches": weight_mismatches,
        "grade_match_rate_on_both": rate(grade_ok),
        "grade_mismatches": grade_mismatches,
        "pcmk_match_rate_on_both": rate(pcmk_ok),
        "pcmk_mismatches": pcmk_mismatches,
    }


def load_reference_bom(path: Path) -> dict[tuple[str, str], int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    agg: dict[tuple[str, str], int] = defaultdict(int)
    for r in rows:
        if not r or all(v is None for v in r[:6]):
            continue
        qty, mat, ln = (list(r) + [None] * 3)[:3]
        try:
            q = int(qty)
        except (TypeError, ValueError):
            continue
        key = (_norm_mat(mat), _norm_len(ln))
        agg[key] += q
    return dict(agg)


def load_generated_bom_from_json(path: Path) -> dict[tuple[str, str], int]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    bom = payload.get("material_summary")
    if not bom or not isinstance(bom, list):
        return {}
    agg: dict[tuple[str, str], int] = defaultdict(int)
    for row in bom:
        if not isinstance(row, dict):
            continue
        try:
            q = int(row.get("qty"))
        except (TypeError, ValueError):
            continue
        key = (_norm_mat(row.get("material")), _norm_len(row.get("length")))
        agg[key] += q
    return dict(agg)


def compare_boms(
    ref: dict[tuple[str, str], int],
    gen: dict[tuple[str, str], int],
) -> dict[str, Any]:
    ref_keys = set(ref)
    gen_keys = set(gen)
    both = ref_keys & gen_keys
    only_ref = ref_keys - gen_keys
    only_gen = gen_keys - ref_keys

    qty_matches = 0
    qty_mismatches: list[tuple[tuple[str, str], int, int]] = []
    for k in both:
        rq, gq = ref[k], gen[k]
        if rq == gq:
            qty_matches += 1
        else:
            qty_mismatches.append((k, rq, gq))

    sum_ref = sum(ref.values())
    sum_gen = sum(gen.values())

    n_ref = len(ref_keys)
    n_gen = len(gen_keys)
    n_both = len(both)

    recall = n_both / n_ref if n_ref else 1.0
    precision = n_both / n_gen if n_gen else 1.0
    f1 = (
        (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0
    )

    qty_accuracy = qty_matches / n_both if n_both else 0.0

    return {
        "reference_line_items": n_ref,
        "generated_line_items": n_gen,
        "keys_in_both": n_both,
        "keys_only_reference": len(only_ref),
        "keys_only_generated": len(only_gen),
        "key_recall_vs_reference": recall,
        "key_precision_vs_generated": precision,
        "key_f1": f1,
        "qty_match_on_intersection": qty_matches,
        "qty_mismatch_on_intersection": len(qty_mismatches),
        "qty_accuracy_on_intersection": qty_accuracy,
        "total_qty_reference": sum_ref,
        "total_qty_generated": sum_gen,
        "total_qty_ratio_gen_over_ref": (sum_gen / sum_ref) if sum_ref else None,
        "only_reference_keys": sorted(only_ref),
        "only_generated_keys": sorted(only_gen),
        "qty_mismatches": qty_mismatches,
    }


def main() -> int:
    import argparse
    import os

    from dotenv import load_dotenv

    parser = argparse.ArgumentParser(
        description="Compare material_summary in takeoff JSON to reference Material Summary xlsx.",
    )
    parser.add_argument(
        "--relaxed-inches",
        type=float,
        default=None,
        help="If set (>0), print relaxed key F1 using this length tolerance in inches "
        "(overrides BOM_RELAXED_LENGTH_INCHES).",
    )
    args, _rest = parser.parse_known_args()

    load_dotenv(".env", override=False)
    ref_path = Path(
        (os.getenv("REFERENCE_BOM_XLSX", "") or "").strip()
        or "26-LQ-094_SADDLEBACK VILLAGE_Material Summary.xlsx"
    ).expanduser()
    json_path = Path(
        (os.getenv("INPUT_JSON", "") or "").strip() or "takeoff_output.json"
    ).expanduser()

    if not ref_path.is_file():
        print(f"ERROR: Reference BOM not found: {ref_path}", file=sys.stderr)
        return 1
    if not json_path.is_file():
        print(f"ERROR: JSON not found: {json_path}", file=sys.stderr)
        return 1

    ref_agg = load_reference_bom(ref_path)
    gen_agg = load_generated_bom_from_json(json_path)
    if not gen_agg:
        print(
            "ERROR: No material_summary in JSON (or empty). "
            "Re-run takeoff with updated prompt so the model emits material_summary.",
            file=sys.stderr,
        )
        return 2

    ref_rows = load_reference_rows(ref_path)
    gen_rows = load_generated_bom_rows(json_path)
    ref_by_mat = rollup_by_material(ref_rows)
    gen_by_mat = rollup_by_material(gen_rows)
    col = compare_columns_by_material(ref_by_mat, gen_by_mat)

    r = compare_boms(ref_agg, gen_agg)

    relaxed_tol = args.relaxed_inches
    if relaxed_tol is None:
        relaxed_tol = float(os.getenv("BOM_RELAXED_LENGTH_INCHES", "0") or "0")

    print("\n=== Column agreement (length IGNORED; values summed per material) ===\n")
    print(
        "Materials in reference only (by normalized material string): "
        f"{col['materials_reference']}"
    )
    print(
        "Materials in generated only: "
        f"{col['materials_generated']}"
    )
    print(f"Materials present in BOTH: {col['materials_in_both']}")
    print(
        f"Material recall (both / ref): {col['material_recall']:.1%}  "
        f"({col['materials_in_both']}/{col['materials_reference']})"
    )
    print(
        f"Material precision (both / gen): {col['material_precision']:.1%}  "
        f"({col['materials_in_both']}/{col['materials_generated']})"
    )
    print(
        f"\nQty — same total qty for that material: {col['qty_match_rate_on_both']:.1%}  "
        f"({col['qty_matches']}/{col['materials_in_both']})"
    )
    wr = col["weight_match_rate_both_numeric"]
    if wr is not None:
        print(
            f"Weight — match where BOTH sides have numeric weight: {wr:.1%}  "
            f"({col['weight_ok_count']}/{col['weight_materials_both_numeric']})"
        )
    else:
        print(
            "Weight — no material had numeric weight on BOTH sides; "
            "compare weight_mismatches for one-sided / missing data."
        )
    print(
        f"Grade — same set of grades (per material): {col['grade_match_rate_on_both']:.1%}  "
        f"on {col['materials_in_both']} materials"
    )
    print(
        f"Pcmk — same set of piece marks (per material): {col['pcmk_match_rate_on_both']:.1%}  "
        f"on {col['materials_in_both']} materials"
    )

    if col["only_reference_materials"]:
        print(
            f"\nReference-only materials ({len(col['only_reference_materials'])}):"
        )
        for m in col["only_reference_materials"][:25]:
            print(f"  {m}")
    if col["only_generated_materials"]:
        print(
            f"\nGenerated-only materials ({len(col['only_generated_materials'])}):"
        )
        for m in col["only_generated_materials"][:25]:
            print(f"  {m}")

    if col["qty_mismatches"]:
        print("\nQty differences (material → ref_qty vs gen_qty):")
        for m, rq, gq in col["qty_mismatches"][:25]:
            print(f"  {m[:60]:<60}  ref={rq}  gen={gq}")
    if col["weight_mismatches"]:
        print("\nWeight differences (or one side null):")
        for item in col["weight_mismatches"][:20]:
            print(f"  {item}")

    print("\n=== BOM accuracy (material × length, normalized) ===\n")
    print(f"Reference file: {ref_path.name}")
    print(f"Generated from: {json_path.name} (material_summary)\n")
    print(f"Unique BOM lines — reference: {r['reference_line_items']}")
    print(f"Unique BOM lines — generated: {r['generated_line_items']}")
    print(f"Keys matching both: {r['keys_in_both']}")
    print(
        f"Key recall (matched / reference): {r['key_recall_vs_reference']:.1%}  "
        f"({r['keys_in_both']}/{r['reference_line_items']})"
    )
    print(
        f"Key precision (matched / generated): {r['key_precision_vs_generated']:.1%}  "
        f"({r['keys_in_both']}/{r['generated_line_items']})"
    )
    print(f"Key F1: {r['key_f1']:.1%}")
    print(
        f"\nWhere keys match — qty exact match: {r['qty_accuracy_on_intersection']:.1%}  "
        f"({r['qty_match_on_intersection']}/{r['keys_in_both']})"
    )
    print(f"Total qty sum — reference: {r['total_qty_reference']}  generated: {r['total_qty_generated']}")
    if r["total_qty_ratio_gen_over_ref"] is not None:
        print(f"Total qty ratio (gen/ref): {r['total_qty_ratio_gen_over_ref']:.3f}")

    if relaxed_tol and relaxed_tol > 0:
        from saddleback_pipeline.bom_relaxed import relaxed_key_match_metrics

        rel = relaxed_key_match_metrics(
            set(ref_agg.keys()),
            set(gen_agg.keys()),
            length_tol_inches=relaxed_tol,
        )
        print(
            f"\n=== Relaxed BOM keys (material + length within ±{relaxed_tol:g}\") ===\n"
        )
        print(
            f"Relaxed key recall: {rel['relaxed_key_recall']:.1%}  "
            f"({rel['reference_keys_with_relaxed_match']}/{rel['reference_keys']})"
        )
        print(
            f"Relaxed key precision: {rel['relaxed_key_precision']:.1%}  "
            f"({rel['generated_keys_with_relaxed_match']}/{rel['generated_keys']})"
        )
        print(f"Relaxed key F1: {rel['relaxed_key_f1']:.1%}")

    print(f"\nLines only in reference ({r['keys_only_reference']}):")
    for k in r["only_reference_keys"][:20]:
        print(f"  {k}  qty={ref_agg[k]}")
    if len(r["only_reference_keys"]) > 20:
        print("  ...")

    print(f"\nLines only in generated ({r['keys_only_generated']}):")
    for k in r["only_generated_keys"][:20]:
        print(f"  {k}  qty={gen_agg[k]}")
    if len(r["only_generated_keys"]) > 20:
        print("  ...")

    if r["qty_mismatches"]:
        print("\nQty mismatches (same key, different qty):")
        for k, rq, gq in r["qty_mismatches"][:25]:
            print(f"  {k}  ref={rq} gen={gq}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
