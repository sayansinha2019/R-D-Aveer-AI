"""Compare a **generated** Project-1 style BOM xlsx to a **reference** BOM (same sheet layout).

Outputs JSON with row counts, category mix, piecemark overlap, and W-shape quantity deltas.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl


def _load_bom_rows(path: Path) -> list[list[Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Bill of Materials"]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 19)]
        if any(v is not None and str(v).strip() for v in row):
            rows.append(row)
    wb.close()
    return rows


def _norm_w_section(sec: Any) -> str | None:
    if sec is None:
        return None
    s = str(sec).strip().upper().replace(" ", "").replace("×", "X")
    m = re.match(r"^W(\d+)X(\d+)$", s)
    if m:
        return f"W{m.group(1)}X{m.group(2)}"
    m = re.match(r"^(\d+)X(\d+)$", s)
    if m:
        return f"W{m.group(1)}X{m.group(2)}"
    return None


def _w_beam_rollup(rows: list[list[Any]]) -> dict[str, int]:
    """Category Beams + Section Type W → sum Qty by normalized W designation."""
    out: dict[str, int] = {}
    for r in rows:
        cat = str(r[1] or "").strip()
        st = str(r[4] or "").strip()
        if cat != "Beams" or st != "W":
            continue
        w = _norm_w_section(r[5])
        if not w:
            continue
        try:
            q = int(r[3])
        except (TypeError, ValueError):
            continue
        out[w] = out.get(w, 0) + q
    return out


def compare_project1_boms(reference_xlsx: Path, generated_xlsx: Path) -> dict[str, Any]:
    ref_rows = _load_bom_rows(reference_xlsx)
    gen_rows = _load_bom_rows(generated_xlsx)

    ref_pm = {r[2] for r in ref_rows if r[2]}
    gen_pm = {r[2] for r in gen_rows if r[2]}

    ref_w = _w_beam_rollup(ref_rows)
    gen_w = _w_beam_rollup(gen_rows)
    common_w = sorted(set(ref_w) & set(gen_w))

    qty_exact = 0
    delta: list[dict[str, Any]] = []
    for w in common_w:
        rq, gq = ref_w[w], gen_w[w]
        if rq == gq:
            qty_exact += 1
        else:
            delta.append(
                {
                    "section": w,
                    "reference_qty": rq,
                    "generated_qty": gq,
                    "ratio_gen_over_ref": round(gq / rq, 4) if rq else None,
                }
            )

    return {
        "reference_path": str(reference_xlsx),
        "generated_path": str(generated_xlsx),
        "row_counts": {
            "reference": len(ref_rows),
            "generated": len(gen_rows),
        },
        "categories_reference": dict(Counter(str(r[1] or "").strip() for r in ref_rows).most_common()),
        "categories_generated": dict(Counter(str(r[1] or "").strip() for r in gen_rows).most_common()),
        "unique_piecemarks": {
            "reference": len(ref_pm),
            "generated": len(gen_pm),
            "intersection": len(ref_pm & gen_pm),
        },
        "w_beam_sections": {
            "reference_distinct": len(ref_w),
            "generated_distinct": len(gen_w),
            "in_both": len(common_w),
        },
        "w_qty_exact_match_on_shared_sections": qty_exact,
        "w_qty_shared_sections_compared": len(common_w),
        "w_qty_match_rate_on_shared": (
            qty_exact / len(common_w) if common_w else None
        ),
        "w_qty_deltas_on_shared": delta,
        "note": "Piecemark IDs differ when the model uses shop marks (B_24) vs generated labels (BEAM-1). "
        "W-shape rollups compare like structural sizes only.",
    }


def main() -> int:
    p = argparse.ArgumentParser(description="Compare Project-1 BOM xlsx to reference.")
    p.add_argument("--reference", type=Path, required=True)
    p.add_argument("--generated", type=Path, required=True)
    p.add_argument("--out-json", type=Path, default=None)
    args = p.parse_args()
    report = compare_project1_boms(args.reference.expanduser().resolve(), args.generated.expanduser().resolve())
    text = json.dumps(report, indent=2, ensure_ascii=False)
    print(text)
    if args.out_json:
        args.out_json.expanduser().resolve().parent.mkdir(parents=True, exist_ok=True)
        args.out_json.write_text(text, encoding="utf-8")
        print(f"Wrote {args.out_json}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
