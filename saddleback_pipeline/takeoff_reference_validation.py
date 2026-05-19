"""Reference-BOM validation helpers for iterative takeoff repair loops."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl

from saddleback_pipeline.project1_bom_export import _classify_row


def load_reference_category_qty(reference_bom_xlsx: Path | str) -> Counter[str]:
    """Return quantity totals by Category from Project-1 BOM."""
    p = Path(reference_bom_xlsx).expanduser().resolve()
    wb = openpyxl.load_workbook(p, data_only=True)
    try:
        ws = wb["Bill of Materials"]
        out: Counter[str] = Counter()
        for r in range(2, ws.max_row + 1):
            cat = str(ws.cell(r, 2).value or "").strip()
            if not cat:
                continue
            qv = ws.cell(r, 4).value
            try:
                q = int(qv)
            except (TypeError, ValueError):
                q = 1
            out[cat] += max(0, q)
        return out
    finally:
        wb.close()


def payload_category_qty(payload: dict[str, Any]) -> Counter[str]:
    """Classify payload entities using export rules and roll up category quantities."""
    out: Counter[str] = Counter()
    for e in payload.get("data") or []:
        if not isinstance(e, dict):
            continue
        category, _section_type, _labor, _is_main = _classify_row(e)
        qv = e.get("quantity")
        try:
            q = int(qv) if qv is not None else 1
        except (TypeError, ValueError):
            q = 1
        out[category] += max(1, q)
    return out
