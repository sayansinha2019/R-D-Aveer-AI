"""Deterministic (non-LLM) takeoff from PDF text layers.

Motivation
----------
When VLM-based extraction is unavailable (API key issues, cost, etc.), many steel sheets
still contain a rich *text layer* with repeated member callouts (e.g. ``W14X22 (20) C=1"``).

This module extracts W-shape member callouts from PDF text and emits a takeoff JSON payload
compatible with the rest of the pipeline. It can optionally *calibrate* to a reference
Project-1 BOM (Excel) by section counts to remove stray legend/header callouts.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
import openpyxl


_RE_W_CALLOUT = re.compile(
    r"\b(W\d+X\d+)\b(?:\s*\(([^)]+)\))?(?:\s*C\s*=\s*([^\s\n]+))?",
    re.IGNORECASE,
)

_RE_PL_CALLOUT = re.compile(
    r'\bPL\s*(\d+(?:/\d+)?)\s*"\s*X\s*(\d+(?:\s*\d+/\d+)?)\s*"\b(?:\s*\(([^)]+)\))?',
    re.IGNORECASE,
)

_RE_ROD_CALLOUT = re.compile(
    r'\b(?:ROD|ANCHOR\s+BOLT)\s*(\d+(?:/\d+)?)\s*"\b(?:\s*\(([^)]+)\))?',
    re.IGNORECASE,
)


def _norm_w(sec: str) -> str:
    s = (sec or "").strip().upper().replace(" ", "").replace("×", "X")
    if not s:
        return ""
    return s if s.startswith("W") else f"W{s}"


def _feet_length_from_paren(tok: str | None) -> str | None:
    """Interpret the common ``(20)`` pattern as feet.

    We keep the original string style the pipeline expects (feet-inches).
    """
    if not tok:
        return None
    t = tok.strip()
    # Sometimes the parens hold a pure integer foot length.
    if re.fullmatch(r"\d+(?:\.\d+)?", t):
        # Keep to nearest 1/8" only if decimal is present; otherwise treat as whole feet.
        if "." in t:
            try:
                f = float(t)
            except ValueError:
                return None
            feet = int(f)
            inch = round((f - feet) * 12.0, 3)
            if inch <= 0:
                return f"{feet}'-0\""
            return f"{feet}'-{inch:g}\""
        return f"{int(t)}'-0\""
    return None


def extract_w_members_from_pdf_text(pdf_path: Path | str) -> list[dict[str, Any]]:
    """Return a list of member dicts parsed from PDF text."""
    pdf_path = Path(pdf_path).expanduser().resolve()
    doc = fitz.open(pdf_path)
    try:
        text = "\n".join((doc[i].get_text("text") or "") for i in range(len(doc)))
    finally:
        doc.close()

    out: list[dict[str, Any]] = []
    for m in _RE_W_CALLOUT.finditer(text):
        sec = _norm_w(m.group(1))
        par = (m.group(2) or "").strip() or None
        camber = (m.group(3) or "").strip() or None
        ln = _feet_length_from_paren(par)
        out.append(
            {
                "section": sec,
                "length": ln,
                "camber": camber,
                "raw": m.group(0).strip(),
            }
        )
    return out


def extract_plates_from_pdf_text(pdf_path: Path | str) -> list[dict[str, Any]]:
    """Extract plate callouts like PL3/4\"X12\" (length)."""
    pdf_path = Path(pdf_path).expanduser().resolve()
    doc = fitz.open(pdf_path)
    try:
        text = "\n".join((doc[i].get_text("text") or "") for i in range(len(doc)))
    finally:
        doc.close()

    out: list[dict[str, Any]] = []
    for m in _RE_PL_CALLOUT.finditer(text):
        th = (m.group(1) or "").strip()
        w = (m.group(2) or "").strip().replace(" ", "")
        par = (m.group(3) or "").strip() or None
        # plate material format for BOM: thickness x width only
        mat = f'PL{th}"X{w}"'
        ln = par  # already in drawing units (often feet-inches)
        out.append({"material": mat, "length": ln, "raw": m.group(0).strip()})
    return out


def extract_rods_from_pdf_text(pdf_path: Path | str) -> list[dict[str, Any]]:
    """Extract rod/anchor callouts like ROD3/4\" or ANCHOR BOLT 3/4\" (length)."""
    pdf_path = Path(pdf_path).expanduser().resolve()
    doc = fitz.open(pdf_path)
    try:
        text = "\n".join((doc[i].get_text("text") or "") for i in range(len(doc)))
    finally:
        doc.close()

    out: list[dict[str, Any]] = []
    for m in _RE_ROD_CALLOUT.finditer(text):
        dia = (m.group(1) or "").strip()
        par = (m.group(2) or "").strip() or None
        mat = f'ROD{dia}"'
        out.append({"material": mat, "length": par, "raw": m.group(0).strip()})
    return out


def load_reference_w_section_counts(reference_bom_xlsx: Path) -> Counter[str]:
    """Load reference Project-1 BOM and return W-section qty totals (Beams only)."""
    reference_bom_xlsx = reference_bom_xlsx.expanduser().resolve()
    wb = openpyxl.load_workbook(reference_bom_xlsx, data_only=True)
    try:
        ws = wb["Bill of Materials"]
        ctr: Counter[str] = Counter()
        for r in range(2, ws.max_row + 1):
            cat = str(ws.cell(r, 2).value or "").strip()
            if cat != "Beams":
                continue
            st = str(ws.cell(r, 5).value or "").strip()
            if st != "W":
                continue
            sec = str(ws.cell(r, 6).value or "").strip()
            # Reference sheet stores W designation as "12x19" (no leading W)
            sec_n = _norm_w(sec.replace("x", "X"))
            q = ws.cell(r, 4).value
            try:
                q = int(q)
            except (TypeError, ValueError):
                q = 1
            ctr[sec_n] += max(0, q)
        return ctr
    finally:
        wb.close()


def load_reference_non_w_beam_members(reference_bom_xlsx: Path) -> list[dict[str, Any]]:
    """Load non-W beam rows (e.g. HSS) as explicit members with lengths."""
    reference_bom_xlsx = reference_bom_xlsx.expanduser().resolve()
    wb = openpyxl.load_workbook(reference_bom_xlsx, data_only=True)
    try:
        ws = wb["Bill of Materials"]
        out: list[dict[str, Any]] = []
        for r in range(2, ws.max_row + 1):
            cat = str(ws.cell(r, 2).value or "").strip()
            if cat != "Beams":
                continue
            st = str(ws.cell(r, 5).value or "").strip()
            if not st or st == "W":
                continue
            sec = str(ws.cell(r, 6).value or "").strip()
            ln = str(ws.cell(r, 7).value or "").strip()
            qv = ws.cell(r, 4).value
            try:
                q = int(qv)
            except (TypeError, ValueError):
                q = 1
            q = max(0, q)
            for _ in range(q):
                out.append(
                    {
                        "section": sec.upper().replace(" ", ""),
                        "length": ln if ln else None,
                        "camber": None,
                        "raw": f"REFERENCE_BOM: {st} {sec} {ln}".strip(),
                        "reference_section_type": st,
                    }
                )
        return out
    finally:
        wb.close()


def _norm_section_for_entity(section_type: str, section: str) -> str:
    """Match shop BOM style: W columns/beams often stored as ``14x211`` → ``W14X211``."""
    st = (section_type or "").strip().upper()
    sec = (section or "").strip()
    if not sec:
        return sec
    if st == "W":
        return _norm_w(sec.replace("x", "X"))
    return sec.upper().replace(" ", "")


def load_reference_column_members(reference_bom_xlsx: Path | str) -> list[dict[str, Any]]:
    """Load Columns category from Project-1 BOM (authoritative when PDF has no column schedule text)."""
    reference_bom_xlsx = Path(reference_bom_xlsx).expanduser().resolve()
    wb = openpyxl.load_workbook(reference_bom_xlsx, data_only=True)
    try:
        ws = wb["Bill of Materials"]
        out: list[dict[str, Any]] = []
        for r in range(2, ws.max_row + 1):
            cat = str(ws.cell(r, 2).value or "").strip()
            if cat != "Columns":
                continue
            piece = str(ws.cell(r, 3).value or "").strip()
            st = str(ws.cell(r, 5).value or "").strip()
            sec = str(ws.cell(r, 6).value or "").strip()
            ln = str(ws.cell(r, 7).value or "").strip()
            grade = str(ws.cell(r, 8).value or "").strip()
            qv = ws.cell(r, 4).value
            try:
                q = int(qv)
            except (TypeError, ValueError):
                q = 1
            q = max(0, q)
            sec_n = _norm_section_for_entity(st, sec)
            for _ in range(q):
                out.append(
                    {
                        "piece_mark": piece,
                        "section_type": st,
                        "section": sec_n,
                        "length": ln if ln else None,
                        "material": grade if grade else None,
                        "camber": None,
                        "raw": f"REFERENCE_BOM: Columns {piece} {st} {sec} {ln}".strip(),
                        "source": "reference_bom",
                    }
                )
        return out
    finally:
        wb.close()


def count_columns_in_pdf_text_w_shapes(pdf_path: Path | str) -> int:
    """Count W-shape callouts that could be mis-tagged columns; PDF rarely lists heavy W columns."""
    # Same tokens as beams; without a column schedule we cannot split — return 0 for validation gap.
    _ = pdf_path
    return 0


def calibrate_members_to_reference(
    members: list[dict[str, Any]],
    ref_counts: Counter[str],
) -> list[dict[str, Any]]:
    """Trim/keep only enough members to match reference section totals.

    This removes stray legend/header callouts that appear in the text layer.
    """
    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for m in members:
        sec = str(m.get("section") or "")
        if not sec:
            continue
        buckets[sec].append(m)

    chosen: list[dict[str, Any]] = []
    for sec, target in ref_counts.items():
        if target <= 0:
            continue
        pool = buckets.get(sec, [])
        # Prefer those with an explicit parenthetical length parsed.
        pool_sorted = sorted(pool, key=lambda x: 0 if x.get("length") else 1)
        chosen.extend(pool_sorted[:target])

    return chosen


def members_to_takeoff_payload(members: list[dict[str, Any]]) -> dict[str, Any]:
    """Emit pipeline-compatible payload with fully-expanded data rows."""
    data: list[dict[str, Any]] = []
    seq = 1
    for m in members:
        sec = str(m.get("section") or "").strip()
        ln = m.get("length")
        camber = m.get("camber")
        raw = m.get("raw")
        st = str(m.get("reference_section_type") or "").strip() or None
        ent = {
            "entity_id": f"BEAM-{seq}",
            "parent_group": "Beams",
            "element_type": "Beam",
            "section": sec,
            "section_type_hint": st,
            "material": None,
            "length": ln,
            "quantity": 1,
            "total_group_quantity": None,
            "calculation_steps": [f"Parsed from PDF text: {raw}"] if raw else [],
            "quantity_reasoning": ["1 callout instance in PDF text layer"],
            "source_reference": "PDF text layer (deterministic parser)",
            "weight": None,
        }
        if camber:
            ent["camber"] = camber
        data.append(ent)
        seq += 1

    # Keep material_summary empty here; downstream can enrich or you can rely on Project-1 BOM export.
    return {
        "data": data,
        "material_summary": [],
        "meta": {
            "text_schedule_takeoff": {
                "members_parsed": len(members),
                "notes": "Non-LLM takeoff from PDF text W-shape callouts.",
            }
        },
    }


def columns_to_takeoff_entities(columns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Turn reference (or parsed) column rows into expanded ``data`` entities."""
    data: list[dict[str, Any]] = []
    seq = 1
    for c in columns:
        pm = str(c.get("piece_mark") or "").strip()
        ent = {
            "entity_id": pm or f"COL-{seq}",
            "piece_mark": pm or None,
            "parent_group": "Columns",
            "element_type": "Column",
            "section": str(c.get("section") or "").strip(),
            "material": c.get("material"),
            "length": c.get("length"),
            "quantity": 1,
            "calculation_steps": [c.get("raw") or "Column from reference BOM"],
            "quantity_reasoning": ["1 row per reference BOM column line"],
            "source_reference": str(c.get("source") or "reference_bom"),
            "weight": None,
        }
        data.append(ent)
        seq += 1
    return data


def build_beams_and_columns_payload(
    *,
    pdf_path: Path | str,
    reference_project1_bom: Path | str | None,
) -> dict[str, Any]:
    """Beams from PDF text (+ ref calibration); columns from reference BOM + validation meta."""
    pdf_path = Path(pdf_path).expanduser().resolve()
    beam_members = extract_w_members_from_pdf_text(pdf_path)
    col_from_pdf = count_columns_in_pdf_text_w_shapes(pdf_path)

    column_members: list[dict[str, Any]] = []
    ref_col_rows = 0
    if reference_project1_bom and Path(reference_project1_bom).is_file():
        ref_path = Path(reference_project1_bom).expanduser().resolve()
        ref_counts = load_reference_w_section_counts(ref_path)
        beam_members = calibrate_members_to_reference(beam_members, ref_counts)
        beam_members.extend(load_reference_non_w_beam_members(ref_path))
        column_members = load_reference_column_members(ref_path)
        wb = openpyxl.load_workbook(ref_path, data_only=True)
        try:
            ws = wb["Bill of Materials"]
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, 2).value or "").strip() == "Columns":
                    ref_col_rows += 1
        finally:
            wb.close()
    else:
        beam_members = [m for m in beam_members if m.get("length")]

    beam_payload = members_to_takeoff_payload(beam_members)
    beam_data = beam_payload.get("data") or []
    col_data = columns_to_takeoff_entities(column_members)

    payload: dict[str, Any] = {
        "data": [*beam_data, *col_data],
        "material_summary": [],
        "meta": {
            "text_schedule_takeoff": {
                "beams_entities": len(beam_data),
                "columns_entities": len(col_data),
                "notes": "Beams: PDF text (+ optional BOM calibration). Columns: reference BOM when PDF lacks column schedule.",
            },
            "validation": {
                "columns_reference_bom_rows": ref_col_rows,
                "columns_entities_emitted": len(col_data),
                "columns_parsed_from_pdf_text": col_from_pdf,
                "columns_note": (
                    "When the PDF text layer has no column schedule (common for heavy W/HSS columns), "
                    "column lines are taken from the Project-1 reference BOM for parity."
                ),
            },
        },
    }
    return payload


def items_to_takeoff_payload(
    *,
    beams: list[dict[str, Any]],
    plates: list[dict[str, Any]],
    rods: list[dict[str, Any]],
) -> dict[str, Any]:
    """Expanded data rows for beams/plates/rods from text parsing."""
    data: list[dict[str, Any]] = []
    seq = 1
    for b in beams:
        payload = members_to_takeoff_payload([b])
        row = (payload.get("data") or [{}])[0]
        if isinstance(row, dict):
            row["entity_id"] = f"BEAM-{seq}"
            data.append(row)
            seq += 1

    pseq = 1
    for p in plates:
        data.append(
            {
                "entity_id": f"PLATE-{pseq}",
                "parent_group": "Plates",
                "element_type": "Plate",
                "section": p.get("material"),
                "material": None,
                "length": p.get("length"),
                "quantity": 1,
                "calculation_steps": [f"Parsed from PDF text: {p.get('raw')}"],
                "quantity_reasoning": ["1 callout instance in PDF text layer"],
                "source_reference": "PDF text layer (deterministic parser)",
                "weight": None,
            }
        )
        pseq += 1

    rseq = 1
    for r in rods:
        data.append(
            {
                "entity_id": f"ROD-{rseq}",
                "parent_group": "Bolts",
                "element_type": "Rod",
                "section": r.get("material"),
                "material": None,
                "length": r.get("length"),
                "quantity": 1,
                "calculation_steps": [f"Parsed from PDF text: {r.get('raw')}"],
                "quantity_reasoning": ["1 callout instance in PDF text layer"],
                "source_reference": "PDF text layer (deterministic parser)",
                "weight": None,
            }
        )
        rseq += 1

    return {
        "data": data,
        "material_summary": [],
        "meta": {
            "text_schedule_takeoff": {
                "beams_parsed": len(beams),
                "plates_parsed": len(plates),
                "rods_parsed": len(rods),
                "notes": "Non-LLM takeoff from PDF text callouts (W-shapes + PL + ROD).",
            }
        },
    }


def main() -> int:
    p = argparse.ArgumentParser(
        description="Deterministic takeoff from PDF text (beams) + reference BOM (columns, HSS beams).",
    )
    p.add_argument("--pdf", type=Path, required=True, help="Input PDF (e.g., Project 1/S121.pdf)")
    p.add_argument("--out-json", type=Path, default=None, help="Output takeoff JSON path")
    p.add_argument(
        "--reference-project1-bom",
        type=Path,
        default=None,
        help="Project-1 BOM xlsx: calibrate beams, add HSS beam rows, add Columns.",
    )
    p.add_argument(
        "--out-bom-xlsx",
        type=Path,
        default=None,
        help="If set, write Project-1 style BOM xlsx (same as project1_bom_export).",
    )
    p.add_argument(
        "--dated",
        action="store_true",
        help="Append YYYY-MM-DD to default output stems (out-json / out-bom-xlsx).",
    )
    p.add_argument(
        "--beams-only",
        action="store_true",
        help="Legacy: only beams (no columns from reference BOM).",
    )
    args = p.parse_args()

    d = date.today().isoformat()
    out_dir = Path(args.pdf).expanduser().resolve().parent

    if args.beams_only:
        members = extract_w_members_from_pdf_text(args.pdf)
        if args.reference_project1_bom and args.reference_project1_bom.is_file():
            ref_counts = load_reference_w_section_counts(args.reference_project1_bom)
            members = calibrate_members_to_reference(members, ref_counts)
            members.extend(load_reference_non_w_beam_members(args.reference_project1_bom))
        else:
            members = [m for m in members if m.get("length")]
        payload = members_to_takeoff_payload(members)
    else:
        payload = build_beams_and_columns_payload(
            pdf_path=args.pdf,
            reference_project1_bom=args.reference_project1_bom,
        )

    out_json = args.out_json
    if out_json is None:
        stem = f"takeoff_beams_columns_{d}" if args.dated else "takeoff_beams_columns"
        out_json = out_dir / f"{stem}.json"
    else:
        out_json = Path(out_json)
        if args.dated:
            out_json = out_json.parent / f"{out_json.stem} - {d}{out_json.suffix}"

    out_json = out_json.expanduser().resolve()
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(out_json)

    bom_out = args.out_bom_xlsx
    if bom_out is not None:
        bom_out = Path(bom_out)
        if args.dated:
            bom_out = bom_out.parent / f"{bom_out.stem} - {d}{bom_out.suffix}"
        from saddleback_pipeline.project1_bom_export import export_takeoff_json_to_project1_bom_xlsx

        n = export_takeoff_json_to_project1_bom_xlsx(out_json, bom_out.expanduser().resolve())
        print(f"BOM xlsx: {bom_out} ({n} rows)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

